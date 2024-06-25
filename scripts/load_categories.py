from openpyxl import load_workbook
from store.models import Category, SubCategory


def run():
  # path = "C:\\pycharm\\django_project\\ctm\\ctm\\utilities\\excel\\categories.xlsx"
  path = "utilities\\excel\\categories.xlsx"
  wb = load_workbook(path)

  ws = wb.active
  m_row = ws.max_row

  for row in range(2, m_row + 1):
    category = ws.cell(row=row, column=1).value
    subcategory = ws.cell(row=row, column=2).value
    
    get_category, created = Category.objects.get_or_create(name=category)
      
    subcat = SubCategory.objects.create(
      name = subcategory,
      category = get_category
    )
    subcat.save()

  
