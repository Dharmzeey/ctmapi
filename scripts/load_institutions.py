from openpyxl import load_workbook
from user.models import Institution, Location, State


def run():
  # path = "C:\\pycharm\\django_project\\ctm\\ctm\\utilities\\excel\\Nigeria_Universities.xlsx"
  path = "utilities\\excel\\Nigeria_Universities.xlsx"
  wb = load_workbook(path)

  ws = wb.active
  m_row = ws.max_row

  for row in range(2, m_row + 1):
    institution = ws.cell(row=row, column=1).value
    location = ws.cell(row=row, column=2).value
    state = ws.cell(row=row, column=3).value
    
    get_state, created = State.objects.get_or_create(name=state)
    get_location, created = Location.objects.get_or_create(name=location, state=get_state)
    
    uni = Institution.objects.create(
      name = institution,
      state = get_state, 
      location = get_location
    )
    uni.save()
  
  
  
